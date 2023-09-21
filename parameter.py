_author__    = "Dr. Pham Quang Phuong","Cao Anh Quoc Hung"
__copyright__ = "Copyright 2023"
__license__   = "All rights reserved"
__email__     = "phuong.phamquang@hust.edu.vn","caoanhquochung@gmail.com"
__status__    = "Released"
__version__   = "1.2.5"
"""
about: ....
"""
import sys,os,time,math
import openpyxl,csv
import argparse
import numpy as np
from scipy.sparse import csr_matrix
"""
import psutil
process = psutil.Process()
print("Memory usage:", process.memory_info().rss)
"""


PARSER_INPUTS = argparse.ArgumentParser(epilog= "")
PARSER_INPUTS.usage = 'Distribution network analysis Tools'
PARSER_INPUTS.add_argument('-fi' , help = '*(str) Input file path' , default = '',type=str,metavar='')
PARSER_INPUTS.add_argument('-fo' , help = '*(str) Output file path', default = '',type=str,metavar='')
ARGVS = PARSER_INPUTS.parse_known_args()[0]
#
RATEC = 100/math.sqrt(3)

#
def toString(v,nRound=5):
    """ convert object/value to String """
    if v is None:
        return 'None'
    t = type(v)
    if t==str:
        if "'" in v:
            return ('"'+v+'"').replace('\n',' ')
        return ("'"+v+"'").replace('\n',' ')
    if t==int:
        return str(v)
    if t==float:
        if v>1.0:
            s1 = str(round(v,nRound))
            return s1[:-2] if s1.endswith('.0') else s1
        elif abs(v)<1e-8:
            return '0'
        s1 ='%.'+str(nRound)+'g'
        return s1 % v
    if t==complex:
        if v.imag>=0:
            return '('+ toString(v.real,nRound)+' +' + toString(v.imag,nRound)+'j)'
        return '('+ toString(v.real,nRound) +' '+ toString(v.imag,nRound)+'j)'
    try:
        return v.toString()
    except:
        pass
    if t in {list,tuple,set}:
        s1=''
        for v1 in v:
            s1+=toString(v1,nRound)+','
        if v:
            s1 = s1[:-1]
        if t==list:
            return '['+s1+']'
        elif t==tuple:
            return '('+s1+')'
        else:
            return '{'+s1+'}'
    if t==dict:
        s1=''
        for k1,v1 in v.items():
            s1+=toString(k1)+':'
            s1+=toString(v1,nRound)+','
        if s1:
            s1 = s1[:-1]
        return '{'+s1+'}'
    return str(v)
#
def add2CSV(nameFile,ares,delim):
    """
    append array String to a file CSV
    """
    pathdir = os.path.split(os.path.abspath(nameFile))[0]
    try:
        os.mkdir(pathdir)
    except:
        pass
    #
    if not os.path.isfile(nameFile):
        with open(nameFile, mode='w') as f:
            ew = csv.writer(f, delimiter=delim, quotechar='"',lineterminator="\n")
            for a1 in ares:
                ew.writerow(a1)
            f.close()
    else:
        with open(nameFile, mode='a') as f:
            ew = csv.writer(f, delimiter=delim, quotechar='"',lineterminator="\n", quoting=csv.QUOTE_MINIMAL)
            for a1 in ares:
                ew.writerow(a1)
#
class Parameter:
    def __init__(self,fi):
        self.tcheck = 0
        wbInput = openpyxl.load_workbook(os.path.abspath(fi),data_only=True)
        loadProfile = self.__readInput1Sheet__(wbInput,'LOADPROFILE')
        genProfile = self.__readInput1Sheet__(wbInput,'GENPROFILE')
        busa = self.__readInput1Sheet__(wbInput,'BUS')
        linea = self.__readInput1Sheet__(wbInput,'LINE')
        dgProfile = self.__readInput1Sheet__(wbInput,'DGPROFILE')
        self.setting = self.__readSetting__(wbInput)
        #
        self.iterMax = int(self.setting['option_PF'][0])
        self.epsilon = self.setting['option_PF'][1]
        self.lineOff = None
        # print(self.setting)
        # BUS[NO] =[ kV,PLOAD[kw],QLOAD[kvar],code ]
        self.BUS = {}
        self.busSlack = []
        self.BUSbs = {} #shunt
        self.Qsh = {}
        self.dg = {}
        for i in range(len(busa['NO'])):
            if busa['FLAG'][i]:
                n1 = busa['NO'][i]
                kv = busa['kV'][i]
                p1 = busa['PLOAD[kw]'][i]/1000
                q1 = busa['QLOAD[kvar]'][i]/1000
                qsh = busa['Qshunt[kvar]'][i]/1000
                pdg = busa['Pdgen[kw]'][i]/1000
                qdg = busa['Qdgen[kvar]'][i]/1000
                if abs(qsh)>1e-6:
                    self.BUSbs[n1] = qsh
                #
                c1 = busa['CODE'][i]
                if c1==None:
                    c1=1
                if c1 in {2,3}:
                    self.busSlack.append(n1)
                #
                self.BUS[n1] = [kv,p1,q1,c1]
                if abs(pdg) > 0 and abs(qdg) > 0: 
                    self.dg[n1] = [pdg,qdg]
        self.nSlack = len(self.busSlack)
        self.setSlack = set(self.busSlack)
        #
        self.Ubase = self.BUS[self.busSlack[0]][0]
        self.Ubase2 = self.Ubase*self.Ubase
        # update B shunt at bus
        for k1,v1 in self.BUSbs.items():
            self.BUSbs[k1] = v1/self.Ubase2# q =u*u*b
        #
        self.profileID = [int(i) for i in loadProfile['time\\NOBUS']]
        self.profiledgID = [int(i) for i in dgProfile['time\\NOBUS']]
        if len(self.profiledgID) != len(self.profiledgID):
            raise Exception('Error size of profile')
        #
        # LOAD PROFILE convert to MVA
        self.loadProfile = dict()
        self.loadAll = dict()
        for ii in range(len(self.profileID)):
            k = loadProfile['time\\NOBUS'][ii]
            v1 = dict()
            self.loadAll[k] = 0
            for k1 in loadProfile.keys():
                if k1!='time\\NOBUS':
                    n1 = int(k1)
                    v1[n1] = loadProfile[k1][ii] * complex(self.BUS[n1][1],self.BUS[n1][2])
                    self.loadAll[k]+=v1[n1]
            self.loadProfile[k] = v1
        # GENPROFILE convert to kV
        self.genProfile = dict()
        for ii in range(len(self.profileID)):
            k = genProfile['time\\NOBUS'][ii]
            v1 = dict()
            for k1 in genProfile.keys():
                if k1!='time\\NOBUS':
                    n1 = int(k1)
                    v1[n1] = genProfile[k1][ii] * self.Ubase
            self.genProfile[k]=v1
        # DG PROFILE convert to MVA
        self.dgProfile = dict()
        self.dgAll = dict()
        if self.dg:
            for ii in range(len(self.profiledgID)):
                k = dgProfile['time\\NOBUS'][ii]
                v1 = dict()
                for k1 in dgProfile.keys():
                    if k1!='time\\NOBUS':
                        n1 = int(k1)
                        v1[n1] = complex(dgProfile[k1][ii] * self.dg[n1][0],dgProfile[k1][ii] *self.dg[n1][1])
                self.dgProfile[k] = v1
        # LINE[NO] = [FROMBUS,TOBUS,RX(Ohm),B/2(Siemens),RATE ]
        self.LINE = {}
        self.LINEb = {} # b of Line
        for i in range(len(linea['NO'])):
            if linea['FLAG'][i]:
                n1= linea['NO'][i]
                fr = linea['FROMBUS'][i]
                to = linea['TOBUS'][i]
                r = linea['R(Ohm)'][i]
                x = linea['X(Ohm)'][i]
                r1 = linea['RATEA[A]'][i]/1000 #kA
                self.LINE[n1] = [fr,to,complex(r,x),r1]
                #
                if linea['B(microSiemens)'][i]>1e-2:
                    self.LINEb[n1] = linea['B(microSiemens)'][i]*1e-6/2
        self.setLineHndAll = set(self.LINE.keys())
        #
        self.setBusHnd = set(self.BUS.keys())
        self.lstBusHnd = busa['NO']
        self.lstLineHnd= linea['NO']
        # list cac line co the dong mo
        self.lineFLAG3 = []
        for i in range(len(linea['NO'])):
            if linea['FLAG3'][i]:
                self.lineFLAG3.append(linea['NO'][i])
        #
        self.shuntFLAG3 = []
        if 'FLAG3' in busa.keys():
            for i in range(len(busa['NO'])):
                if busa['FLAG3'][i]:
                    self.shuntFLAG3.append(busa['NO'][i])
        #
        self.dgFLAG3 = []
        self.dgOn = []
        if 'FLAG4' in busa.keys():
            for i in range(len(busa['NO'])):
                if busa['FLAG4'][i]:
                    self.dgFLAG3.append(busa['NO'][i])
                    self.dgOn.append(busa['NO'][i])
        if self.dgOn:
            for ii in range(len(self.profiledgID)):
                k = dgProfile['time\\NOBUS'][ii]
                v1 = dict()
                self.dgAll[k] = 0
                for k1 in self.dgOn:
                    if k1!='time\\NOBUS':
                        n1 = int(k1)
                        v1[n1] = complex(dgProfile[str(k1)][ii] * self.dg[n1][0],dgProfile[str(k1)][ii] *self.dg[n1][1])
                        self.dgAll[k] += v1[n1]
        #
        self.nL = len(self.lineFLAG3)
        self.nSht = len(self.shuntFLAG3)
        self.nDg = len(self.dgFLAG3)
        self.nVar = self.nL + self.nSht + self.nDg
        #
         #
        self.BUSC = dict() #connect of BUS
        for b1 in self.setBusHnd:
            self.BUSC[b1] = set()
        #
        for k,v in self.LINE.items():
            self.BUSC[v[0]].add(k)
            self.BUSC[v[1]].add(k)
        self.AllLine2Bus = {k:self.LINE[k][:2] for k in self.LINE.keys()}
        self.busesin1line = [set(v) for v in self.AllLine2Bus.values()]
        self.listline = [k for k in self.AllLine2Bus.keys()]
        self.bus2bus = {k:set() for k in self.setBusHnd}
        for v in self.LINE.values():
            self.bus2bus[v[0]].add(v[1])
            self.bus2bus[v[1]].add(v[0])
        #
        self.t0 = time.time()
        #print(self.BUSC)
        #print(self.lineISL) # line ko the off, off=>island
        #assume that only take circumstances where there is no island bus to calculate power flow
        self.nBus = len(self.setBusHnd)
        #print('busISL',self.busISL)   # bus con lai sau khi da bo line island, dung de check loop
    def __readInput1Sheet__(self,wbInput,sh1):
        ws = wbInput[sh1]
        res = {}
        # dem so dong data
        for i in range(2,20000):
            if ws.cell(i,1).value==None:
                k=i
                break
        #
        for i in range(1,20000):
            v1 = ws.cell(2,i).value
            if v1==None:
                return res
            va = []
            #
            for i1 in range(3,k):
                va.append(ws.cell(i1,i).value)
            res[str(v1)]=va
        return res
    #
    def __readSetting__(self,wbInput):
        ws = wbInput['SETTING']
        k = 0
        res = {}
        while True:
            k+=1
            s1= ws.cell(k,1).value
            if type(s1)==str and s1.replace(' ','')=='##BRANCHING':
                for j in range(1,100):
                    s2 = str(ws.cell(k+1,j).value).strip()
                    if s2=='None':
                        break
                    sa = str(ws.cell(k+2,j).value).split(',')
                    if len(sa)==1:
                        try:
                            res[s2] = float(sa[0])
                        except:
                            res[s2] = sa[0]
                    else:
                        res[s2] = [float(si) for si in sa]
                break
        return res
    #